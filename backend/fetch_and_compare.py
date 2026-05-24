"""
Pull project '2025' classifications from the live API and compare vs manual corrections.
Usage:  PYTHONUTF8=1 python fetch_and_compare.py
"""

import re
import sys
import requests
import pandas as pd
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

API_BASE   = "https://calc-carbon-140293665526.me-west1.run.app"
PROJECT    = "2025-8"
MANUAL_FILE = r"C:\Users\user\OneDrive - post.bgu.ac.il\מסמכים\יזמות\זיהוי שגוי.xlsx"
OUTPUT_FILE = r"C:\Users\user\OneDrive - post.bgu.ac.il\מסמכים\יזמות\2025-בדיקת-דיוק-חדש.xlsx"

# ── helpers ───────────────────────────────────────────────────────────────────

def norm_text(t) -> str:
    if pd.isna(t): return ""
    return re.sub(r"\s+", " ", str(t).strip().lower())

def strip_parens(cat) -> str:
    if pd.isna(cat): return ""
    return re.sub(r"\s*\(.*?\)", "", str(cat)).strip()

_NO_CARBON = {"EXCLUDE", "Unknown", "None", ""}

def cats_match(shira_cat, correct_cat) -> bool:
    s = strip_parens(shira_cat) if not pd.isna(shira_cat) else ""
    c = strip_parens(correct_cat) if not pd.isna(correct_cat) else ""
    if not s: s = "EXCLUDE"
    if not c: c = "EXCLUDE"
    if s == c: return True
    if s in _NO_CARBON and c in _NO_CARBON: return True
    return False

# ── fetch from API ────────────────────────────────────────────────────────────

print(f"Fetching project '{PROJECT}' from API…")
resp = requests.get(f"{API_BASE}/emissions", params={"project": PROJECT, "limit": 50000}, timeout=60)
resp.raise_for_status()
items = resp.json()["items"]
print(f"  Got {len(items)} rows")

shira = pd.DataFrame(items)
# Normalise column names to match comparison script expectations
col_map = {
    "boq_code":         "קוד BOQ",
    "short_text":       "תיאור",
    "category":         "קטגוריה",
    "matched_by":       "שיטה",
    "reliability_score":"ביטחון %",
}
shira = shira.rename(columns=col_map)
# Fill missing description from 'material' if needed
if "תיאור" not in shira.columns and "material" in shira.columns:
    shira["תיאור"] = shira["material"]

print(f"  Columns: {list(shira.columns)}")

# ── load manual corrections ───────────────────────────────────────────────────

print("Loading manual corrections…")
manual = pd.read_excel(MANUAL_FILE, sheet_name="ניסיון 1", header=1)
manual.columns = [c.strip() for c in manual.columns]
print(f"  Manual: {len(manual)} rows")

shira["_norm"]   = shira["תיאור"].apply(norm_text)
manual["_norm"]  = manual["תיאור"].apply(norm_text)
manual["_cat_norm"] = manual["קטגוריה נכונה"].apply(strip_parens)

# Build lookup
lookup_exact: dict = {}
lookup_prefix: dict = {}
for _, row in manual.iterrows():
    k   = row["_norm"]
    cat = row["_cat_norm"]
    if not k or not cat:
        continue
    if k not in lookup_exact:
        lookup_exact[k] = cat
    k40 = k[:40]
    if k40 and k40 not in lookup_prefix:
        lookup_prefix[k40] = cat

print(f"  Lookup: {len(lookup_exact)} exact keys")

# ── match & classify ──────────────────────────────────────────────────────────

matched_correct, matched_incorrect, unmatched = [], [], []

for _, row in shira.iterrows():
    norm      = row.get("_norm", "")
    shira_cat = str(row.get("קטגוריה", "")).strip()

    correct_cat = lookup_exact.get(norm)
    match_type  = "exact"
    if correct_cat is None:
        correct_cat = lookup_prefix.get(norm[:40])
        match_type  = "prefix40"

    if correct_cat is None:
        unmatched.append({
            "קוד BOQ":       row.get("קוד BOQ", ""),
            "תיאור":         row.get("תיאור", ""),
            "קטגוריה SHIRA": shira_cat,
            "שיטה":          row.get("שיטה", ""),
            "ביטחון %":      row.get("ביטחון %", ""),
        })
        continue

    ok = cats_match(shira_cat, correct_cat)
    record = {
        "קוד BOQ":              row.get("קוד BOQ", ""),
        "תיאור":                row.get("תיאור", ""),
        "קטגוריה SHIRA":        shira_cat,
        "קטגוריה נכונה (ידני)": correct_cat,
        "שיטה":                 row.get("שיטה", ""),
        "ביטחון %":             row.get("ביטחון %", ""),
        "_match_type":          match_type,
    }
    (matched_correct if ok else matched_incorrect).append(record)

total   = len(shira)
n_ok    = len(matched_correct)
n_wrong = len(matched_incorrect)
n_none  = len(unmatched)
pct     = 100 * n_ok / (n_ok + n_wrong) if (n_ok + n_wrong) else 0

print(f"\n{'─'*50}")
print(f"Total rows        : {total}")
print(f"Matched to manual : {n_ok + n_wrong}  ({n_none} unmatched)")
print(f"✓ Correct         : {n_ok}  ({pct:.1f}%)")
print(f"✗ Wrong           : {n_wrong}")
print(f"{'─'*50}")

# ── per-category summary ──────────────────────────────────────────────────────

cat_counts: dict = defaultdict(lambda: {"correct": 0, "wrong": 0})
for r in matched_correct:   cat_counts[r["קטגוריה נכונה (ידני)"]]["correct"] += 1
for r in matched_incorrect: cat_counts[r["קטגוריה נכונה (ידני)"]]["wrong"]   += 1

summary_rows = sorted([
    {"קטגוריה": cat, "נכון": d["correct"], "שגוי": d["wrong"],
     'סה"כ': d["correct"]+d["wrong"],
     "דיוק %": round(100*d["correct"]/(d["correct"]+d["wrong"]), 1) if d["correct"]+d["wrong"] else 0}
    for cat, d in cat_counts.items()
], key=lambda x: -x["שגוי"])

# ── Excel output ──────────────────────────────────────────────────────────────

RED    = PatternFill("solid", fgColor="FFCCCC")
GREEN  = PatternFill("solid", fgColor="CCFFCC")
YELLOW = PatternFill("solid", fgColor="FFFFCC")
HEADER = PatternFill("solid", fgColor="2E7D32")
HDR_FONT = Font(color="FFFFFF", bold=True)

def write_sheet(ws, rows, fill, title):
    if not rows:
        ws.append(["(אין נתונים)"]); ws.title = title; return
    df = pd.DataFrame(rows)
    df = df[[c for c in df.columns if not c.startswith("_")]]
    ws.append(list(df.columns))
    for cell in ws[1]:
        cell.fill = HEADER; cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center")
    for row in df.itertuples(index=False):
        ws.append(list(row))
        for cell in ws[ws.max_row]: cell.fill = fill
    for i, col in enumerate(df.columns, 1):
        mx = max(len(str(col)), df[col].astype(str).str.len().max() if len(df) else 0)
        ws.column_dimensions[get_column_letter(i)].width = min(mx + 4, 60)
    ws.title = title

wb = Workbook()
wb.remove(wb.active)

write_sheet(wb.create_sheet(), matched_incorrect, RED,    f"חשוד - לא תואם ({n_wrong})")
write_sheet(wb.create_sheet(), matched_correct,   GREEN,  f"תואם - נכון ({n_ok})")
write_sheet(wb.create_sheet(), unmatched,         YELLOW, f"לא הותאם ({n_none})")

ws4 = wb.create_sheet(title="סיכום")
ws4.append(["קטגוריה", "נכון", "שגוי", 'סה"כ', "דיוק %"])
for cell in ws4[1]:
    cell.fill = HEADER; cell.font = HDR_FONT
    cell.alignment = Alignment(horizontal="center")
for r in summary_rows:
    ws4.append([r["קטגוריה"], r["נכון"], r["שגוי"], r['סה"כ'], r["דיוק %"]])
    fill = GREEN if r["דיוק %"] >= 80 else (YELLOW if r["דיוק %"] >= 50 else RED)
    for cell in ws4[ws4.max_row]: cell.fill = fill
ws4.append(['סה"כ', n_ok, n_wrong, n_ok+n_wrong, round(pct, 1)])
for cell in ws4[ws4.max_row]: cell.font = Font(bold=True)
for i, w in enumerate([40, 10, 10, 10, 12], 1):
    ws4.column_dimensions[get_column_letter(i)].width = w

wb.save(OUTPUT_FILE)
print(f"\nSaved → {OUTPUT_FILE}")
